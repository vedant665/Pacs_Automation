"""
Excel Report Generator for PACS Automation Test Suite
Generates a 4-sheet Excel report: Summary, Test Guide, Details, Screenshots

Module-agnostic engine with injectable descriptions & categories.
FP (Forgot Password) is the built-in default.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


# ─── Color Palette ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="666666")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SKIP_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUMMARY_BG = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


# ─── Default Descriptions (Forgot Password — Login Module) ───────────────────
# Keys are matched against test function names (partial match supported)
# Format: short human-readable question anyone can understand

TEST_DESCRIPTIONS = {
    # ── Screen 1: Email Entry ──
    "test_fp_s1_01": {
        "name": "Unregistered Email",
        "question": "Does the app still send an OTP even if the email isn't registered? (This is by design to prevent leaking who has an account)",
    },
    "test_fp_s1_02": {
        "name": "Valid Email",
        "question": "Does a registered email successfully trigger an OTP to be sent?",
    },
    "test_fp_s1_03": {
        "name": "Blank Email",
        "question": "Does the app stop you if you click 'Send OTP' without typing any email?",
    },
    "test_fp_s1_04": {
        "name": "Email with Spaces",
        "question": "Does it handle accidental spaces like '  vedant@email.com  ' correctly?",
    },
    "test_fp_s1_05": {
        "name": "Case Insensitive Email",
        "question": "Does VEDANT@EMAIL.COM work the same as vedant@email.com?",
    },
    "test_fp_s1_06": {
        "name": "Double-Click Send OTP",
        "question": "What happens if you accidentally click 'Send OTP' twice — does it break or send duplicates?",
    },

    # ── Screen 2: OTP + Password Entry ──
    "test_fp_s2_01": {
        "name": "Wrong OTP",
        "question": "Does the app block incorrect OTP codes and show an error message?",
    },
    "test_fp_s2_02": {
        "name": "Password Mismatch",
        "question": "Does it catch you if the 'New Password' and 'Confirm Password' don't match?",
    },
    "test_fp_s2_03": {
        "name": "Weak Password",
        "question": "Does the app enforce strong password rules (minimum length, uppercase, number, special character)?",
    },
    "test_fp_s2_04": {
        "name": "Old Password Reuse",
        "question": "Can you set a password you've used before? (You shouldn't be able to — last 4 passwords are blocked)",
    },

    # ── Full Flow Tests ──
    "test_fp_ff_04": {
        "name": "Full Reset + Login",
        "question": "THE BIG ONE: Does the entire forgot password work end-to-end — reset password, log in with new password, and land on the dashboard?",
    },
    "test_fp_ff_01": {
        "name": "Recently Used Password Blocked",
        "question": "Does the app reject the password you JUST changed from? (Prevents cycling between two passwords)",
    },
    "test_fp_ff_02": {
        "name": "Back to Login Link",
        "question": "Does clicking 'Back to Login' take you back to the login page correctly?",
    },
    "test_fp_ff_03": {
        "name": "Current Password Blocked",
        "question": "Can you 'reset' to the exact same password you're currently using? (You shouldn't be able to)",
    },
    "test_fp_ff_05": {
        "name": "Browser Back Button",
        "question": "Can someone cheat by hitting the browser Back button after entering OTP to reuse the form?",
    },
}


# ─── Default Category Map (Forgot Password — Login Module) ───────────────────
# Maps test class names to category strings for the Summary sheet

DEFAULT_CATEGORIES = {
    "TestForgotPasswordScreen1": "Screen 1 - Email Entry",
    "TestForgotPasswordScreen2": "Screen 2 - OTP & Password",
    "TestForgotPasswordFullFlow": "Full Flow",
}


# ─── Helper Functions ────────────────────────────────────────────────────────

def _get_test_info(nodeid: str, descriptions: dict = None) -> dict:
    """Extract function name from pytest nodeid and find matching description."""
    desc_map = descriptions if descriptions is not None else TEST_DESCRIPTIONS
    func_name = nodeid.split("::")[-1] if "::" in nodeid else nodeid

    if func_name in desc_map:
        return desc_map[func_name]

    # Partial match fallback — match by test ID prefix (e.g. "test_fp_s1_03")
    func_lower = func_name.lower()
    for key, desc in desc_map.items():
        if func_lower.startswith(key):
            return desc

    # Final fallback for unknown tests
    return {
        "name": func_name.replace("test_", "").replace("_", " ").title(),
        "question": "Automated validation step in the test flow",
    }


def _get_category(nodeid: str, category_map: dict = None) -> str:
    """Extract category from test class name."""
    cat_map = category_map if category_map is not None else DEFAULT_CATEGORIES

    if "::" not in nodeid:
        return "Unknown"
    parts = nodeid.split("::")
    class_name = parts[-2] if len(parts) >= 3 else ""

    return cat_map.get(class_name, "Unknown")


def _status_fill(status: str) -> PatternFill:
    if status == "PASSED":
        return PASS_FILL
    elif status == "FAILED":
        return FAIL_FILL
    elif status == "ERROR":
        return ERROR_FILL
    return SKIP_FILL


def _status_label(status: str) -> str:
    if status == "PASSED":
        return "PASS"
    elif status == "FAILED":
        return "FAIL"
    elif status == "ERROR":
        return "ERROR"
    return "SKIP"


def _clean_error_message(message: str) -> str:
    """Clean up pytest error messages for non-technical readers."""
    if not message:
        return ""

    # Remove file paths and line numbers
    import re
    cleaned = re.sub(r'File ".*?", line \d+', '', message)
    cleaned = re.sub(r'tests[/\\].*\.py', '', cleaned)
    cleaned = re.sub(r'\n', ' ', cleaned)

    # Remove excessive whitespace
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    # Truncate if too long
    if len(cleaned) > 200:
        cleaned = cleaned[:197] + "..."

    return cleaned


def generate_report(results: list, output_dir: str = "reports",
                    title: str = None, filename_prefix: str = "ForgotPassword",
                    descriptions: dict = None, categories: dict = None) -> str:
    """
    Generate Excel report with 4 sheets.

    Args:
        results: List of dicts with keys: nodeid, status, message, duration, screenshot
        output_dir: Directory to save the report
        title: Report title shown on sheets (default: "Forgot Password")
        filename_prefix: Prefix for output filename (default: "ForgotPassword")
        descriptions: Dict of test descriptions (default: FP descriptions)
        categories: Dict mapping class_name -> category (default: FP categories)

    Returns:
        Path to the generated Excel file
    """
    # Resolve defaults
    report_title = title or "Forgot Password"

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_TestReport_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    # ─────────────────────────────────────────────────────────────────────
    # SHEET 1: Summary
    # ─────────────────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "1F4E79"

    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"{report_title} - Test Execution Summary"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    ws_summary.merge_cells("A2:F2")
    ts_cell = ws_summary["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
    ts_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    ts_cell.alignment = Alignment(horizontal="center")

    summary_headers = ["Category", "Total Tests", "Passed", "Failed", "Error", "Pass Rate"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Build category stats — dynamic (works for any module)
    cat_stats = {}
    for r in results:
        cat = _get_category(r["nodeid"], categories)
        if cat not in cat_stats:
            cat_stats[cat] = {"total": 0, "passed": 0, "failed": 0, "error": 0}
        cat_stats[cat]["total"] += 1
        status = r["status"].upper()
        if status == "PASSED":
            cat_stats[cat]["passed"] += 1
        elif status == "FAILED":
            cat_stats[cat]["failed"] += 1
        elif status == "ERROR":
            cat_stats[cat]["error"] += 1

    row = 5
    total_all = {"total": 0, "passed": 0, "failed": 0, "error": 0}
    for cat_name in cat_stats:
        stats = cat_stats[cat_name]
        rate = (stats["passed"] / stats["total"] * 100) if stats["total"] > 0 else 0
        ws_summary.cell(row=row, column=1, value=cat_name).border = THIN_BORDER
        ws_summary.cell(row=row, column=2, value=stats["total"]).border = THIN_BORDER
        c3 = ws_summary.cell(row=row, column=3, value=stats["passed"])
        c3.border = THIN_BORDER
        if stats["passed"] == stats["total"]:
            c3.fill = PASS_FILL
        c4 = ws_summary.cell(row=row, column=4, value=stats["failed"])
        c4.border = THIN_BORDER
        if stats["failed"] > 0:
            c4.fill = FAIL_FILL
        c5 = ws_summary.cell(row=row, column=5, value=stats["error"])
        c5.border = THIN_BORDER
        if stats["error"] > 0:
            c5.fill = ERROR_FILL
        rate_cell = ws_summary.cell(row=row, column=6, value=f"{rate:.0f}%")
        rate_cell.border = THIN_BORDER
        rate_cell.alignment = CENTER_ALIGN
        if rate == 100:
            rate_cell.fill = PASS_FILL
        elif rate < 50:
            rate_cell.fill = FAIL_FILL
        for c in range(1, 7):
            ws_summary.cell(row=row, column=c).font = BODY_FONT
            ws_summary.cell(row=row, column=c).alignment = CENTER_ALIGN if c > 1 else Alignment(vertical="center")
        row += 1
        for k in total_all:
            total_all[k] += stats[k]

    total_rate = (total_all["passed"] / total_all["total"] * 100) if total_all["total"] > 0 else 0
    for col_idx, val in enumerate(
        ["TOTAL", total_all["total"], total_all["passed"], total_all["failed"], total_all["error"], f"{total_rate:.0f}%"], 1
    ):
        cell = ws_summary.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN if col_idx > 1 else Alignment(vertical="center")
    ws_summary.cell(row=row, column=1).fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    ws_summary.column_dimensions["A"].width = 35
    for col in ["B", "C", "D", "E", "F"]:
        ws_summary.column_dimensions[col].width = 14

    # ─────────────────────────────────────────────────────────────────────
    # SHEET 2: Test Guide (for non-technical users)
    # ─────────────────────────────────────────────────────────────────────
    ws_guide = wb.create_sheet("Test Guide", 1)
    ws_guide.sheet_properties.tabColor = "2E75B6"

    # Title
    ws_guide.merge_cells("A1:D1")
    g_title = ws_guide["A1"]
    g_title.value = f"{report_title} - What We Tested"
    g_title.font = TITLE_FONT
    g_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 35

    # Summary line
    passed_count = sum(1 for r in results if r["status"].upper() == "PASSED")
    failed_count = sum(1 for r in results if r["status"].upper() == "FAILED")
    error_count = sum(1 for r in results if r["status"].upper() == "ERROR")
    total_count = len(results)
    date_str = datetime.now().strftime("%d-%b-%Y")

    if failed_count == 0 and error_count == 0:
        summary_text = f"{total_count} out of {total_count} tests passed | {date_str}"
    else:
        summary_text = f"{passed_count} out of {total_count} tests passed | {failed_count} failed | {date_str}"

    ws_guide.merge_cells("A2:D2")
    g_summary = ws_guide["A2"]
    g_summary.value = summary_text
    g_summary.font = Font(name="Calibri", size=11, bold=True,
                          color="006100" if failed_count == 0 else "9C0006")
    g_summary.alignment = Alignment(horizontal="center")
    ws_guide.row_dimensions[2].height = 25

    # Headers — 4 columns now
    guide_headers = ["Test", "What Was Checked", "Result", "Why?"]
    for col_idx, header in enumerate(guide_headers, 1):
        cell = ws_guide.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws_guide.row_dimensions[4].height = 25

    # Build result and message lookups
    result_lookup = {}
    message_lookup = {}
    for r in results:
        func_name = r["nodeid"].split("::")[-1] if "::" in r["nodeid"] else r["nodeid"]
        result_lookup[func_name] = r["status"].upper()
        message_lookup[func_name] = r.get("message", "")

    # Collect unique tests in order
    seen_funcs = set()
    ordered_tests = []
    for r in results:
        func_name = r["nodeid"].split("::")[-1] if "::" in r["nodeid"] else r["nodeid"]
        if func_name not in seen_funcs:
            seen_funcs.add(func_name)
            ordered_tests.append(r)

    for idx, r in enumerate(ordered_tests):
        func_name = r["nodeid"].split("::")[-1] if "::" in r["nodeid"] else r["nodeid"]
        info = _get_test_info(r["nodeid"], descriptions)
        actual = result_lookup.get(func_name, "NOT RUN")
        raw_message = message_lookup.get(func_name, "")

        # Determine the "Why?" text based on status
        if actual == "PASSED":
            why_text = "App behaved correctly"
        elif actual == "FAILED":
            why_text = _clean_error_message(raw_message) if raw_message else "Test failed — check Details sheet for more info"
        elif actual == "ERROR":
            why_text = _clean_error_message(raw_message) if raw_message else "Test encountered an error — check Details sheet"
        else:
            why_text = _clean_error_message(raw_message) if raw_message else "Test was skipped"

        # Column A: Test name
        cell_a = ws_guide.cell(row=5 + idx, column=1, value=info["name"])
        cell_a.font = Font(name="Calibri", size=10, bold=True)
        cell_a.alignment = Alignment(vertical="center")
        cell_a.border = THIN_BORDER

        # Column B: What was checked (the question)
        cell_b = ws_guide.cell(row=5 + idx, column=2, value=info["question"])
        cell_b.font = BODY_FONT
        cell_b.alignment = WRAP_ALIGN
        cell_b.border = THIN_BORDER

        # Column C: Result
        cell_c = ws_guide.cell(row=5 + idx, column=3, value=_status_label(actual))
        cell_c.font = Font(name="Calibri", size=10, bold=True)
        cell_c.alignment = CENTER_ALIGN
        cell_c.fill = _status_fill(actual)
        cell_c.border = THIN_BORDER

        # Column D: Why?
        cell_d = ws_guide.cell(row=5 + idx, column=4, value=why_text)
        cell_d.font = BODY_FONT
        cell_d.alignment = WRAP_ALIGN
        cell_d.border = THIN_BORDER

        # If failed/error, tint the "Why?" cell red/yellow
        if actual == "FAILED":
            cell_d.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        elif actual == "ERROR":
            cell_d.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws_guide.row_dimensions[5 + idx].height = 35

    # Column widths
    ws_guide.column_dimensions["A"].width = 28
    ws_guide.column_dimensions["B"].width = 60
    ws_guide.column_dimensions["C"].width = 10
    ws_guide.column_dimensions["D"].width = 50

    # ─────────────────────────────────────────────────────────────────────
    # SHEET 3: Details
    # ─────────────────────────────────────────────────────────────────────
    ws_details = wb.create_sheet("Details")
    ws_details.sheet_properties.tabColor = "548235"

    detail_headers = ["#", "Category", "Test Method", "Status", "Error Message", "Duration (s)", "Timestamp"]
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for idx, r in enumerate(results, 1):
        cat = _get_category(r["nodeid"], categories)
        func_name = r["nodeid"].split("::")[-1] if "::" in r["nodeid"] else r["nodeid"]
        status = r["status"].upper()
        duration = r.get("duration", 0)
        timestamp = r.get("timestamp", "")

        row_data = [idx, cat, func_name, status, r.get("message", ""), f"{duration:.2f}", timestamp]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_details.cell(row=1 + idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_idx == 4:
                cell.alignment = CENTER_ALIGN
                cell.fill = _status_fill(status)
            elif col_idx == 6:
                cell.alignment = CENTER_ALIGN
            elif col_idx == 5:
                cell.alignment = WRAP_ALIGN

    ws_details.column_dimensions["A"].width = 5
    ws_details.column_dimensions["B"].width = 30
    ws_details.column_dimensions["C"].width = 45
    ws_details.column_dimensions["D"].width = 12
    ws_details.column_dimensions["E"].width = 50
    ws_details.column_dimensions["F"].width = 14
    ws_details.column_dimensions["G"].width = 20

    # ─────────────────────────────────────────────────────────────────────
    # SHEET 4: Screenshots
    # ─────────────────────────────────────────────────────────────────────
    ws_shots = wb.create_sheet("Screenshots")
    ws_shots.sheet_properties.tabColor = "BF8F00"

    shot_headers = ["#", "Category", "Test Method", "Screenshot Path"]
    for col_idx, header in enumerate(shot_headers, 1):
        cell = ws_shots.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    shot_idx = 0
    for r in results:
        if r.get("screenshot"):
            shot_idx += 1
            cat = _get_category(r["nodeid"], categories)
            func_name = r["nodeid"].split("::")[-1] if "::" in r["nodeid"] else r["nodeid"]
            row_data = [shot_idx, cat, func_name, r["screenshot"]]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_shots.cell(row=1 + shot_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER

    ws_shots.column_dimensions["A"].width = 5
    ws_shots.column_dimensions["B"].width = 30
    ws_shots.column_dimensions["C"].width = 45
    ws_shots.column_dimensions["D"].width = 60

    # ─────────────────────────────────────────────────────────────────────
    # Save
    # ─────────────────────────────────────────────────────────────────────
    wb.save(filepath)
    return filepath