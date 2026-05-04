"""
validation_report_generator.py
------------------------------
Generates Excel report for Company Onboarding validation test results.
Uses openpyxl for formatting.

Output:
  - Summary sheet: totals, pass/fail/bug counts
  - Test Results sheet: every test with field, bad value, expected, actual
  - Bug Report sheet: only tests that found real bugs
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from common.logger import log


PASS_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
FAIL_FILL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
BUG_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECTION_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_validation_report(results, output_dir):
    """Generate Excel report from validation test results."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"CO_Validation_Report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ─── Sheet 1: Summary ───
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 5
    ws_summary.column_dimensions["B"].width = 40
    ws_summary.column_dimensions["C"].width = 20

    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASSED")
    failed = sum(1 for r in results if r["status"] == "FAILED")
    bugs = sum(1 for r in results if r.get("is_bug"))

    # Title row
    ws_summary.merge_cells("A1:C1")
    cell = ws_summary["A1"]
    cell.value = "Company Onboarding - Validation Test Report"
    cell.font = TITLE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

    # Meta info
    meta = [
        ("", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Bugs Found", bugs),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws_summary[f"A{i}"] = ""
        ws_summary[f"B{i}"] = label
        ws_summary[f"C{i}"] = value
        ws_summary[f"B{i}"].font = Font(bold=True, size=11)
        if label == "Bugs Found" and bugs > 0:
            ws_summary[f"C{i}"].font = Font(bold=True, size=11, color="E74C3C")
        elif label == "Passed":
            ws_summary[f"C{i}"].font = Font(bold=True, size=11, color="27AE60")
        elif label == "Failed":
            ws_summary[f"C{i}"].font = Font(bold=True, size=11, color="E74C3C")

    # ─── Sheet 2: Test Results ───
    ws_results = wb.create_sheet("Test Results")
    columns = [
        ("#", 5),
        ("Test Name", 40),
        ("Category", 15),
        ("Field", 20),
        ("Bad Value", 30),
        ("Expected", 30),
        ("Actual", 30),
        ("Status", 12),
        ("Screenshot", 40),
    ]
    for i, (name, width) in enumerate(columns, 1):
        ws_results.cell(row=1, column=i, value=name)
        ws_results.cell(row=1, column=i).font = HEADER_FONT
        ws_results.cell(row=1, column=i).fill = HEADER_FILL
        ws_results.cell(row=1, column=i).alignment = CENTER
        ws_results.column_dimensions[chr(64 + i)].width = width

    for idx, r in enumerate(results, 2):
        is_bug = r.get("is_bug", False)
        ws_results.cell(row=idx, column=1, value=idx - 1)
        ws_results.cell(row=idx, column=2, value=r.get("test_name", ""))
        ws_results.cell(row=idx, column=3, value=r.get("category", ""))
        ws_results.cell(row=idx, column=4, value=r.get("field", ""))
        ws_results.cell(row=idx, column=5, value=str(r.get("bad_value", ""))[:50])
        ws_results.cell(row=idx, column=6, value=r.get("expected", ""))
        ws_results.cell(row=idx, column=7, value=r.get("actual", ""))
        status = "BUG" if is_bug else r.get("status", "")
        ws_results.cell(row=idx, column=8, value=status)
        ws_results.cell(row=idx, column=9, value=r.get("screenshot", ""))

        fill = BUG_FILL if is_bug else (PASS_FILL if status == "PASSED" else FAIL_FILL)
        font = WHITE_FONT
        ws_results.cell(row=idx, column=8).fill = fill
        ws_results.cell(row=idx, column=8).font = font
        ws_results.cell(row=idx, column=8).alignment = CENTER

        for col in range(1, 10):
            ws_results.cell(row=idx, column=col).border = THIN_BORDER
            ws_results.cell(row=idx, column=col).alignment = LEFT_WRAP

    # ─── Sheet 3: Bug Report (only bugs) ───
    bug_results = [r for r in results if r.get("is_bug")]
    if bug_results:
        ws_bugs = wb.create_sheet("Bug Report")
        bug_columns = [
            ("#", 5),
            ("Test Name", 40),
            ("Field", 20),
            ("Expected Behavior", 40),
            ("Actual Behavior", 40),
            ("Impact", 40),
            ("Screenshot", 40),
        ]
        for i, (name, width) in enumerate(bug_columns, 1):
            ws_bugs.cell(row=1, column=i, value=name)
            ws_bugs.cell(row=1, column=i).font = HEADER_FONT
            ws_bugs.cell(row=1, column=i).fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            ws_bugs.cell(row=1, column=i).alignment = CENTER
            ws_bugs.column_dimensions[chr(64 + i)].width = width

        for idx, r in enumerate(bug_results, 2):
            ws_bugs.cell(row=idx, column=1, value=idx - 1)
            ws_bugs.cell(row=idx, column=2, value=r.get("test_name", ""))
            ws_bugs.cell(row=idx, column=3, value=r.get("field", ""))
            ws_bugs.cell(row=idx, column=4, value=r.get("expected", ""))
            ws_bugs.cell(row=idx, column=5, value=r.get("actual", ""))
            ws_bugs.cell(row=idx, column=6, value=r.get("impact", ""))
            ws_bugs.cell(row=idx, column=7, value=r.get("screenshot", ""))
            for col in range(1, 8):
                ws_bugs.cell(row=idx, column=col).border = THIN_BORDER
                ws_bugs.cell(row=idx, column=col).alignment = LEFT_WRAP

    wb.save(filepath)
    log.info(f"Validation report saved: {filepath}")
    return filepath