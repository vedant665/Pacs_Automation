"""
co_update_report_generator.py
Excel report for Company Onboarding UPDATE tests.
3 Sheets: Field Verification, Summary, All Fields Before vs After
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_DARK  = PatternFill("solid", fgColor="1F3864")
FILL_MED   = PatternFill("solid", fgColor="2E75B6")
FILL_LIGHT = PatternFill("solid", fgColor="D6E4F0")
FILL_ALT   = PatternFill("solid", fgColor="F5F5F5")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_RED   = PatternFill("solid", fgColor="FFC7CE")
FILL_ORANGE = PatternFill("solid", fgColor="FDE9D9")

FONT_TITLE    = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_SUBTITLE = Font(name="Calibri", italic=True, color="FFFFFF", size=11)
FONT_HEADER   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_NORMAL   = Font(name="Calibri", size=10)
FONT_BOLD     = Font(name="Calibri", bold=True, size=10)
FONT_KPI_VAL  = Font(name="Calibri", bold=True, size=18, color="1F3864")
FONT_PASS     = Font(name="Calibri", bold=True, size=10, color="006100")
FONT_FAIL     = Font(name="Calibri", bold=True, size=10, color="9C0006")
FONT_CHANGED  = Font(name="Calibri", bold=True, size=10, color="C65911")

BORDER = Border(
    left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"),
)
A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

STEP_LABELS = {
    "step1": "Step 1: Company Details", "step2": "Step 2: Promoters",
    "step3": "Step 3: Address", "step4": "Step 4: Business Details",
    "step5": "Step 5: Infrastructure",
}
FIELD_LABELS = {
    "company_name": "Company Name", "company_short_name": "Short Name",
    "contact_name": "Contact Name", "company_background": "Company Background",
    "email": "Email", "mobile_number": "Mobile Number",
    "pan": "PAN", "gstin": "GSTIN", "cin": "CIN",
    "entity_group": "Entity Group", "parent_name": "Parent Name", "plan_type": "Plan Type",
    "promoter_name": "Promoter Name", "promoter_remark": "Promoter Remark",
    "address_type": "Address Type", "country": "Country",
    "state": "State", "district": "District", "taluka": "Taluka",
    "address": "Address", "pin_code": "Pin Code",
    "business_model": "Business Model", "market_linkages": "Market Linkages",
    "line_of_business": "Line of Business", "additional_business_activities": "Additional Business",
    "infra_type": "Infrastructure Type", "infra_location": "Infrastructure Location",
    "ownership_type": "Ownership Type",
}
ALL_FIELDS = {
    "step1": ["company_name","company_short_name","contact_name","company_background",
              "email","mobile_number","pan","gstin","cin","entity_group","parent_name","plan_type"],
    "step2": ["promoter_name","promoter_remark"],
    "step3": ["address_type","country","state","district","taluka","address","pin_code"],
    "step4": ["business_model","market_linkages","line_of_business","additional_business_activities"],
    "step5": ["infra_type","infra_location","ownership_type"],
}
UPDATED_FIELDS = {
    1: ["contact_name","email","mobile_number"],
    2: ["promoter_name","promoter_remark"],
    3: ["address","pin_code"],
    4: ["business_model","market_linkages"],
    5: ["infra_location"],
}

def _sc(cell, font=None, fill=None, align=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border

def _get(data, step_key, field, row_idx="1"):
    step = data.get(step_key, {})
    if isinstance(step, dict):
        r = step.get(row_idx, step.get(1, {}))
        if isinstance(r, dict):
            return r.get(field, "")
    return ""

def generate_co_update_report(update_results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fp = os.path.join(output_dir, f"CompanyOnboarding_UpdateReport_{ts}.xlsx")
    wb = Workbook()
    _build_field_verification(wb, update_results)
    _build_summary(wb, update_results)
    _build_all_fields(wb, update_results)
    wb.save(fp)
    return fp

def _build_field_verification(wb, results):
    ws = wb.active
    ws.title = "Field Verification"
    ws.sheet_properties.tabColor = "1F3864"
    r = wb.create_sheet("Summary")
    wb.move_sheet(r, offset=-1)
    ws = wb["Field Verification"]
    for res in results:
        company = res.get("company_name", "")
        ts = res.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.merge_cells("A1:G1")
        _sc(ws["A1"], FONT_TITLE, FILL_DARK, A_C)
        ws["A1"].value = f"UPDATE REPORT: {company}"
        ws.row_dimensions[1].height = 36
        ws.merge_cells("A2:G2")
        _sc(ws["A2"], FONT_SUBTITLE, FILL_MED, A_C)
        ws["A2"].value = f"Generated: {ts}"
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 8
        row = 4
        for i, h in enumerate(["#","Step","Field","Before (Original)","Updated To","After (Read Back)","Status"], 1):
            _sc(ws.cell(row=row, column=i, value=h), FONT_HEADER, FILL_MED, A_C, BORDER)
        ws.row_dimensions[row].height = 24
        row += 1
        before = res.get("before", {})
        after = res.get("after", {})
        updates = res.get("updates_applied", {})
        idx = 0
        for step_num in range(1, 6):
            step_key = f"step{step_num}"
            step_label = STEP_LABELS[step_key]
            step_updates = updates.get(step_num, {})
            ufields = UPDATED_FIELDS[step_num]
            for uf in ufields:
                row_data = step_updates.get("1", step_updates.get(1, step_updates))
                if not isinstance(row_data, dict):
                    row_data = step_updates if isinstance(step_updates, dict) else {}
                updated_val = str(row_data.get(uf, ""))
                if not updated_val:
                    continue
                idx += 1
                before_val = _get(before, step_key, uf)
                after_val = _get(after, step_key, uf)
                status = "PASS" if str(updated_val).strip() == str(after_val).strip() else "FAIL"
                alt = idx % 2 == 0
                fill = FILL_ALT if alt else FILL_WHITE
                vals = [idx, step_label, FIELD_LABELS.get(uf, uf), before_val, updated_val, after_val, status]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=v)
                    if ci == 7:
                        if status == "PASS":
                            _sc(c, FONT_PASS, FILL_GREEN, A_C, BORDER)
                        else:
                            _sc(c, FONT_FAIL, FILL_RED, A_C, BORDER)
                    else:
                        _sc(c, FONT_NORMAL, fill, A_L if ci > 1 else A_L, BORDER)
                row += 1
        for col, w in {"A":10,"B":28,"C":24,"D":35,"E":35,"F":35,"G":12}.items():
            ws.column_dimensions[col].width = w
        break

def _build_summary(wb, results):
    ws = wb["Summary"]
    ws.sheet_properties.tabColor = "2E75B6"
    res = results[0] if results else {}
    company = res.get("company_name", "")
    ws.merge_cells("A1:K1")
    _sc(ws["A1"], FONT_TITLE, FILL_DARK, A_C)
    ws["A1"].value = "UPDATE TEST SUMMARY"
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:K2")
    _sc(ws["A2"], FONT_SUBTITLE, FILL_MED, A_C)
    ws["A2"].value = f"Company: {company}"
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8
    updates = res.get("updates_applied", {})
    before = res.get("before", {})
    after = res.get("after", {})
    total = 0
    passed = 0
    step_stats = {}
    for step_num in range(1, 6):
        step_key = f"step{step_num}"
        step_label = STEP_LABELS[step_key]
        step_updates = updates.get(step_num, {})
        row_data = step_updates.get("1", step_updates.get(1, step_updates))
        if not isinstance(row_data, dict):
            row_data = step_updates if isinstance(step_updates, dict) else {}
        ufields = UPDATED_FIELDS[step_num]
        sp = 0
        sf = 0
        for uf in ufields:
            uv = str(row_data.get(uf, ""))
            if not uv:
                continue
            total += 1
            av = _get(after, step_key, uf)
            if uv.strip() == av.strip():
                sp += 1
                passed += 1
            else:
                sf += 1
        step_stats[step_num] = {"label": step_label, "total": len(ufields), "passed": sp, "failed": sf}
    failed = total - passed
    rate = (passed / total * 100) if total else 0
    for i, (label, val, fill) in enumerate([
        ("Total Fields Updated", str(total), FILL_LIGHT),
        ("Passed", str(passed), FILL_GREEN),
        ("Failed", str(failed), FILL_RED),
        ("Pass Rate", f"{rate:.1f}%", FILL_GREEN if rate == 100 else FILL_RED if rate == 0 else PatternFill("solid", fgColor="FFEB9C")),
    ]):
        col = i * 3 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        _sc(ws.cell(row=4, column=col, value=label), FONT_BOLD, FILL_MED, A_C, BORDER)
        ws.cell(row=4, column=col + 1).border = BORDER
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        _sc(ws.cell(row=5, column=col, value=val), FONT_KPI_VAL, fill, A_C, BORDER)
        ws.cell(row=5, column=col + 1).border = BORDER
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 36
    row = 8
    for ci, h in enumerate(["#","Step","Fields Updated","Passed","Failed","Status"], 1):
        _sc(ws.cell(row=row, column=ci, value=h), FONT_HEADER, FILL_MED, A_C, BORDER)
    ws.row_dimensions[row].height = 24
    row += 1
    for step_num in range(1, 6):
        st = step_stats[step_num]
        alt = step_num % 2 == 0
        fill = FILL_ALT if alt else FILL_WHITE
        status = "PASS" if st["failed"] == 0 else "FAIL"
        vals = [step_num, st["label"], st["total"], st["passed"], st["failed"], status]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            if ci == 6:
                if status == "PASS":
                    _sc(c, FONT_PASS, FILL_GREEN, A_C, BORDER)
                else:
                    _sc(c, FONT_FAIL, FILL_RED, A_C, BORDER)
            else:
                _sc(c, FONT_NORMAL, fill, A_L, BORDER)
        row += 1
    for col, w in {"A":5,"B":28,"C":16,"D":10,"E":10,"F":12}.items():
        ws.column_dimensions[col].width = w

def _build_all_fields(wb, results):
    ws = wb.create_sheet("All Fields Before vs After")
    ws.sheet_properties.tabColor = "C65911"
    res = results[0] if results else {}
    company = res.get("company_name", "")
    ws.merge_cells("A1:F1")
    _sc(ws["A1"], FONT_TITLE, FILL_DARK, A_C)
    ws["A1"].value = "ALL FIELDS - BEFORE vs AFTER"
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:F2")
    _sc(ws["A2"], FONT_SUBTITLE, FILL_MED, A_C)
    ws["A2"].value = f"Company: {company} | Showing all readable fields"
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8
    row = 4
    for ci, h in enumerate(["#","Step","Field","Before","After","Changed"], 1):
        _sc(ws.cell(row=row, column=ci, value=h), FONT_HEADER, FILL_MED, A_C, BORDER)
    ws.row_dimensions[row].height = 24
    row += 1
    before = res.get("before", {})
    after = res.get("after", {})
    idx = 0
    for step_num in range(1, 6):
        step_key = f"step{step_num}"
        step_label = STEP_LABELS[step_key]
        fields = ALL_FIELDS[step_key]
        for field in fields:
            idx += 1
            alt = idx % 2 == 0
            fill = FILL_ALT if alt else FILL_WHITE
            bval = _get(before, step_key, field)
            aval = _get(after, step_key, field)
            changed = "YES" if str(bval) != str(aval) else "No"
            vals = [idx, step_label, FIELD_LABELS.get(field, field), bval, aval, changed]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                if ci == 6 and changed == "YES":
                    _sc(c, FONT_CHANGED, FILL_ORANGE, A_C, BORDER)
                else:
                    _sc(c, FONT_NORMAL, fill, A_L, BORDER)
            row += 1
    for col, w in {"A":5,"B":28,"C":24,"D":35,"E":35,"F":12}.items():
        ws.column_dimensions[col].width = w