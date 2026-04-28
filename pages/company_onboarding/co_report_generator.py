"""
co_report_generator.py
---------------------
Data-focused Excel report for Company Onboarding.

6 Sheets: Summary, Company Details, Promoters, Addresses, Business, Infrastructure
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color Palette ────────────────────────────────────────────
FILL_DARK  = PatternFill("solid", fgColor="1F3864")
FILL_MED   = PatternFill("solid", fgColor="2E75B6")
FILL_LIGHT = PatternFill("solid", fgColor="D6E4F0")
FILL_ALT   = PatternFill("solid", fgColor="F5F5F5")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_RED   = PatternFill("solid", fgColor="FFC7CE")
FILL_GOLD  = PatternFill("solid", fgColor="FFEB9C")

FONT_TITLE    = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_SUBTITLE = Font(name="Calibri", italic=True, color="FFFFFF", size=11)
FONT_HEADER   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_NORMAL   = Font(name="Calibri", size=10)
FONT_BOLD     = Font(name="Calibri", bold=True, size=10)
FONT_KPI_VAL  = Font(name="Calibri", bold=True, size=18, color="1F3864")
FONT_PASS     = Font(name="Calibri", bold=True, size=10, color="006100")
FONT_FAIL     = Font(name="Calibri", bold=True, size=10, color="9C0006")

BORDER = Border(
    left=Side("thin", color="D9D9D9"),
    right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"),
    bottom=Side("thin", color="D9D9D9"),
)

A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _sc(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border


def _title_banner(ws, title, subtitle, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    _sc(c, FONT_TITLE, FILL_DARK, A_C)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c = ws.cell(row=2, column=1, value=subtitle)
    _sc(c, FONT_SUBTITLE, FILL_MED, A_C)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8


def _headers(ws, row, heads, start=1):
    for i, h in enumerate(heads, start):
        c = ws.cell(row=row, column=i, value=h)
        _sc(c, FONT_HEADER, FILL_MED, A_C, BORDER)
    ws.row_dimensions[row].height = 24
    return row + 1


def _row(ws, row, vals, start=1, alt=False):
    fill = FILL_ALT if alt else FILL_WHITE
    for i, v in enumerate(vals, start):
        c = ws.cell(row=row, column=i, value=v)
        _sc(c, FONT_NORMAL, fill, A_L, BORDER)
    return row + 1


def _status_cell(ws, row, col, status):
    c = ws.cell(row=row, column=col, value=status)
    if status == "PASSED":
        _sc(c, FONT_PASS, FILL_GREEN, A_C, BORDER)
    else:
        _sc(c, FONT_FAIL, FILL_RED, A_C, BORDER)


def _widths(ws, spec, mn=10, mx=45):
    for col, w in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(w, mn), mx)


def generate_co_report(companies, results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fp = os.path.join(output_dir, f"CompanyOnboarding_DataReport_{ts}.xlsx")
    wb = Workbook()
    _build_summary(wb, companies, results)
    _build_company(wb, companies, results)
    _build_promoters(wb, companies)
    _build_addresses(wb, companies)
    _build_business(wb, companies)
    _build_infra(wb, companies)
    wb.save(fp)
    return fp


def _build_summary(wb, companies, results):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F3864"
    total = len(companies)
    passed = sum(1 for r in results if r.get("status") == "PASSED")
    failed = total - passed
    rate = (passed / total * 100) if total else 0
    _title_banner(ws, "COMPANY ONBOARDING \u2014 DATA REPORT",
                 f"Generated: {datetime.now().strftime('')}")
    r = 4
    kpis = [
        ("Total Companies", str(total), FILL_LIGHT),
        ("Passed", str(passed), FILL_GREEN),
        ("Failed", str(failed), FILL_RED),
        ("Pass Rate", f"{rate:.1f}%", FILL_GREEN if rate == 100 else FILL_GOLD if 0 < rate < 100 else FILL_RED),
    ]
    for i, (label, val, fill) in enumerate(kpis):
        col = i * 3 + 1
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        _sc(ws.cell(row=r, column=col, value=label), FONT_BOLD, FILL_MED, A_C, BORDER)
        ws.cell(row=r, column=col + 1).border = BORDER
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        _sc(ws.cell(row=r + 1, column=col, value=val), FONT_KPI_VAL, fill, A_C, BORDER)
        ws.cell(row=r + 1, column=col + 1).border = BORDER
        ws.row_dimensions[r + 1].height = 36
    ws.row_dimensions[r].height = 24
    r = 8
    r = _headers(ws, r, ["#", "Company Name", "Short Name", "Entity Group", "Status", "Duration (s)", "Timestamp"])
    for i, (comp, res) in enumerate(zip(companies, results), 1):
        alt = i % 2 == 0
        r = _row(ws, r, [i, comp.get("company_name", ""), comp.get("company_short_name", ""),
                         comp.get("entity_group", ""), res.get("status", ""),
                         f"{res.get('duration', 0):.1f}", res.get("timestamp", "")], alt=alt)
        _status_cell(ws, r - 1, 5, res.get("status", ""))
    _widths(ws, {1: 5, 2: 30, 3: 15, 4: 15, 5: 12, 6: 14, 7: 20})


def _build_company(wb, companies, results):
    ws = wb.create_sheet("Company Details")
    ws.sheet_properties.tabColor = "2E75B6"
    _title_banner(ws, "COMPANY DETAILS", "All header fields per company", 15)
    heads = ["#", "Company Name", "Short Name", "Contact", "Email", "Mobile",
             "PAN", "GSTIN", "CIN", "Entity Group", "Parent", "Linked", "2FA", "Auth Type", "Status"]
    r = _headers(ws, 4, heads)
    for i, (comp, res) in enumerate(zip(companies, results), 1):
        linked = comp.get("company_linked", [])
        if isinstance(linked, list):
            linked = ", ".join(linked)
        vals = [i, comp.get("company_name", ""), comp.get("company_short_name", ""),
                comp.get("contact_name", ""), comp.get("email", ""), comp.get("mobile_number", ""),
                comp.get("pan", ""), comp.get("gstin", ""), comp.get("cin", ""),
                comp.get("entity_group", ""), comp.get("parent_name", ""), linked,
                "Yes" if comp.get("is_2fa") else "No", comp.get("auth_type", ""),
                res.get("status", "")]
        alt = i % 2 == 0
        r = _row(ws, r, vals, alt=alt)
        _status_cell(ws, r - 1, 15, res.get("status", ""))
    _widths(ws, {1: 5, 2: 28, 3: 14, 4: 16, 5: 26, 6: 14, 7: 14, 8: 18, 9: 22,
               10: 14, 11: 14, 12: 14, 13: 6, 14: 12, 15: 10})


def _build_promoters(wb, companies):
    ws = wb.create_sheet("Promoters")
    ws.sheet_properties.tabColor = "548235"
    _title_banner(ws, "PROMOTERS", "Promoter details per company")
    r = _headers(ws, 4, ["#", "Company Name", "#", "Name", "Remark"])
    idx = 1
    for comp in companies:
        for j, p in enumerate(comp.get("promoters", []), 1):
            r = _row(ws, r, [idx, comp.get("company_name", ""), j,
                            p.get("name", ""), p.get("remark", "")], alt=idx % 2 == 0)
            idx += 1
    _widths(ws, {1: 5, 2: 28, 3: 5, 4: 30, 5: 55})


def _build_addresses(wb, companies):
    ws = wb.create_sheet("Addresses")
    ws.sheet_properties.tabColor = "BF8F00"
    _title_banner(ws, "ADDRESSES", "Address details per company", 10)
    r = _headers(ws, 4, ["#", "Company Name", "#", "Type", "Country",
                          "State", "District", "Taluka", "Address", "PIN"])
    idx = 1
    for comp in companies:
        for j, a in enumerate(comp.get("addresses", []), 1):
            r = _row(ws, r, [idx, comp.get("company_name", ""), j,
                            a.get("address_type", ""), a.get("country", ""),
                            a.get("state", ""), a.get("district", ""),
                            a.get("taluka", ""), a.get("address", ""),
                            a.get("pin_code", "")], alt=idx % 2 == 0)
            idx += 1
    _widths(ws, {1: 5, 2: 28, 3: 5, 4: 18, 5: 10, 6: 16, 7: 16, 8: 16, 9: 30, 10: 12})


def _build_business(wb, companies):
    ws = wb.create_sheet("Business Details")
    ws.sheet_properties.tabColor = "7030A0"
    _title_banner(ws, "BUSINESS DETAILS", "Business activity details per company", 7)
    r = _headers(ws, 4, ["#", "Company Name", "#", "Business Model",
                          "Market Linkages", "Line of Business", "Additional Activities"])
    idx = 1
    for comp in companies:
        for j, b in enumerate(comp.get("business_details", []), 1):
            r = _row(ws, r, [idx, comp.get("company_name", ""), j,
                            b.get("business_model", ""), b.get("market_linkages", ""),
                            b.get("line_of_business", ""), b.get("additional_business_activities", "")],
                    alt=idx % 2 == 0)
            idx += 1
    _widths(ws, {1: 5, 2: 28, 3: 5, 4: 35, 5: 35, 6: 35, 7: 45})


def _build_infra(wb, companies):
    ws = wb.create_sheet("Infrastructure")
    ws.sheet_properties.tabColor = "C00000"
    _title_banner(ws, "INFRASTRUCTURE", "Infrastructure details per company", 6)
    r = _headers(ws, 4, ["#", "Company Name", "#", "Type", "Location", "Ownership"])
    idx = 1
    for comp in companies:
        for j, inf in enumerate(comp.get("infrastructure", []), 1):
            r = _row(ws, r, [idx, comp.get("company_name", ""), j,
                            inf.get("infra_type", ""), inf.get("infra_location", ""),
                            inf.get("ownership_type", "")], alt=idx % 2 == 0)
            idx += 1
    _widths(ws, {1: 5, 2: 28, 3: 5, 4: 22, 5: 30, 6: 18})
